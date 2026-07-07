"""Excel test-data reader, built on openpyxl. This is the framework's primary
data-driven source: each sheet's header row becomes dict keys."""
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def read_excel(relative_path: str, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """relative_path is relative to the testdata/excel directory, e.g. 'login.xlsx'.

    Returns a list of row-dicts keyed by the first row's column headers.
    Blank trailing rows are skipped.
    """
    base = Path(__file__).resolve().parent.parent / "testdata" / "excel"
    path = base / relative_path
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    rows_iter = sheet.iter_rows(values_only=True)
    headers = [str(h).strip() for h in next(rows_iter)]

    records: list[dict[str, Any]] = []
    for row in rows_iter:
        if row is None or all(cell is None for cell in row):
            continue
        record = dict(zip(headers, row))
        records.append(record)

    workbook.close()
    return records
